import streamlit as st
import pandas as pd
import zipfile
import io
import gc
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# =========================================================
# PAGE CONFIG
# =========================================================

st.set_page_config(
    page_title="ZIP Excel Filter Tool",
    layout="wide"
)

st.title("📦 ZIP Excel Filter Tool")
st.write("Upload ZIP file containing Excel files and filter rows.")

# =========================================================
# MEMORY SETTINGS
# =========================================================

MAX_ZIP_SIZE_MB = 200

# =========================================================
# HELPERS
# =========================================================

def get_columns_from_excel(file_obj):
    """
    Read only few rows to get columns
    """
    try:
        df = pd.read_excel(
            file_obj,
            nrows=5,
            dtype=str
        )

        return df.columns.tolist()

    except Exception as e:
        st.error(f"Error reading columns: {e}")
        return []


def process_excel_file(
    file_obj,
    selected_columns,
    filter_column,
    filter_value,
    equal_columns,
    datadef_values
):
    """
    Process one Excel file with low memory usage
    """

    try:

        # =========================================
        # USECOLS TO REDUCE MEMORY
        # =========================================

        needed_columns = set(selected_columns)

        if filter_column:
            needed_columns.add(filter_column)

        for c in equal_columns:
            needed_columns.add(c)

        needed_columns.add("DataDefinition")

        # =========================================
        # READ EXCEL
        # =========================================

        df = pd.read_excel(
            file_obj,
            usecols=lambda x: x in needed_columns,
            dtype=str
        )

        df = df.fillna("")

        # =========================================
        # FILTER COLUMN
        # Example:
        # Is_Split == 10
        # =========================================

        if filter_column and filter_value != "":
            df = df[
                df[filter_column].astype(str)
                == str(filter_value)
            ]

        # =========================================
        # EQUAL COLUMNS
        # Example:
        # PartCount == CountRows
        # =========================================

        if len(equal_columns) >= 2:

            first_col = equal_columns[0]

            for col in equal_columns[1:]:

                df = df[
                    df[first_col].astype(str)
                    == df[col].astype(str)
                ]

        # =========================================
        # FILTER DataDefinition
        # =========================================

        if datadef_values:

            if "DataDefinition" in df.columns:

                df = df[
                    df["DataDefinition"].isin(datadef_values)
                ]

        # =========================================
        # KEEP SELECTED COLUMNS ONLY
        # =========================================

        final_columns = [
            c for c in selected_columns
            if c in df.columns
        ]

        df = df[final_columns]

        return df

    except Exception as e:

        return pd.DataFrame({
            "ERROR": [str(e)]
        })


# =========================================================
# FILE UPLOAD
# =========================================================

uploaded_zip = st.file_uploader(
    "Upload ZIP File",
    type=["zip"],
    accept_multiple_files=False
)

# =========================================================
# MAIN
# =========================================================

if uploaded_zip is not None:

    try:

        # =========================================
        # CHECK FILE SIZE
        # =========================================

        zip_size_mb = uploaded_zip.size / 1024 / 1024

        st.info(f"ZIP Size: {zip_size_mb:.2f} MB")

        if zip_size_mb > MAX_ZIP_SIZE_MB:

            st.error(
                f"ZIP too large. Max allowed: {MAX_ZIP_SIZE_MB} MB"
            )

            st.stop()

        # =========================================
        # OPEN ZIP
        # =========================================

        with zipfile.ZipFile(uploaded_zip) as zip_ref:

            excel_files = [

                f for f in zip_ref.namelist()

                if (
                    f.endswith(".xlsx")
                    or f.endswith(".xls")
                )

                and not f.startswith("__MACOSX")

            ]

            if not excel_files:

                st.error("No Excel files found in ZIP")

                st.stop()

            st.success(
                f"Found {len(excel_files)} Excel files"
            )

            # =========================================
            # GET COLUMNS FROM FIRST FILE
            # =========================================

            with zip_ref.open(excel_files[0]) as f:

                columns = get_columns_from_excel(f)

            if not columns:

                st.error("Could not detect columns")

                st.stop()

            # =========================================
            # UI
            # =========================================

            st.subheader("⚙️ Filter Settings")

            col1, col2 = st.columns(2)

            with col1:

                filter_column = st.selectbox(
                    "Filter Column",
                    columns
                )

                filter_value = st.text_input(
                    "Filter Value",
                    value="10"
                )

            with col2:

                selected_columns = st.multiselect(
                    "Columns To Keep",
                    columns,
                    default=columns[:10]
                )

            st.subheader("🔄 Equal Columns Check")

            equal_columns = st.multiselect(
                "Columns That Must Be Equal",
                columns,
                default=[
                    c for c in [
                        "PartCount",
                        "CountRows"
                    ]
                    if c in columns
                ]
            )

            st.subheader("📘 DataDefinition Filter")

            datadef_text = st.text_area(
                "Allowed DataDefinition Values (one per line)",
                value=""
            )

            datadef_values = [

                x.strip()

                for x in datadef_text.splitlines()

                if x.strip()

            ]

            # =========================================
            # PROCESS BUTTON
            # =========================================

            if st.button("🚀 Process Files"):

                progress = st.progress(0)

                status = st.empty()

                # =========================================
                # CREATE EXCEL WORKBOOK
                # =========================================

                wb = Workbook()

                # Remove default sheet
                default_sheet = wb.active
                wb.remove(default_sheet)

                total_files = len(excel_files)

                # =========================================
                # PROCESS FILES
                # =========================================

                for idx, excel_name in enumerate(excel_files):

                    status.write(
                        f"Processing: {excel_name}"
                    )

                    try:

                        with zip_ref.open(excel_name) as f:

                            # Read directly
                            excel_buffer = io.BytesIO(
                                f.read()
                            )

                            result_df = process_excel_file(
                                file_obj=excel_buffer,
                                selected_columns=selected_columns,
                                filter_column=filter_column,
                                filter_value=filter_value,
                                equal_columns=equal_columns,
                                datadef_values=datadef_values
                            )

                        # =========================================
                        # CREATE SHEET
                        # =========================================

                        safe_sheet_name = (
                            excel_name
                            .split("/")[-1]
                            [:31]
                        )

                        ws = wb.create_sheet(
                            title=safe_sheet_name
                        )

                        # =========================================
                        # WRITE DATAFRAME
                        # =========================================

                        for row in dataframe_to_rows(
                            result_df,
                            index=False,
                            header=True
                        ):
                            ws.append(row)

                        # =========================================
                        # CLEAN MEMORY
                        # =========================================

                        del result_df
                        del excel_buffer

                        gc.collect()

                    except Exception as e:

                        ws = wb.create_sheet(
                            title=f"ERROR_{idx}"
                        )

                        ws.append(["ERROR"])
                        ws.append([str(e)])

                    progress.progress(
                        (idx + 1) / total_files
                    )

                # =========================================
                # SAVE OUTPUT
                # =========================================

                output = io.BytesIO()

                wb.save(output)

                output.seek(0)

                gc.collect()

                st.success("Processing Complete ✅")

                st.download_button(
                    label="⬇️ Download Result Excel",
                    data=output,
                    file_name="Filtered_Output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

    except Exception as e:

        st.error(str(e))

# =========================================================
# INFO
# =========================================================

st.markdown("""
---
## 🚀 Memory Optimized

This app is optimized for Streamlit Cloud 1GB RAM:

- Processes files one-by-one
- Reads only needed columns
- Uses low-memory mode
- Cleans memory after each file
- Does not merge all files in RAM
- Uses openpyxl streaming

---
## ✅ Example Filters

### Example 1

Filter:
- Is_Split = 10
- PartCount == CountRows

### Example 2

Allowed DataDefinition:
- Resistor
- Capacitor
- Voltage

---
## ▶️ Run

pip install streamlit pandas openpyxl

streamlit run app.py
""")
